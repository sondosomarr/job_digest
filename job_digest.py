#!/usr/bin/env python3
"""
Daily job digest for Sondos Omar Ragab Hashem.

What this does, every time it runs:
  1. Pulls fresh postings from public, scrape-friendly sources:
       - Wuzzuf's public RSS feed (all Egypt jobs)
       - RemoteOK's public JSON API (remote jobs worldwide)
     Nothing here logs into any account or touches LinkedIn/Indeed -
     both explicitly forbid automated access in their Terms of Service.
  2. Filters for DevOps / Cloud / Platform / SRE roles at junior or
     entry level (senior/lead/principal/staff/manager titles are dropped).
  3. Scores each posting against the skills on the CV, so the sheet
     is sorted with the best matches on top.
  4. Merges the results into tracker.xlsx, keeping any "Applied"
     status you've already set - it only appends rows it hasn't seen
     before (matched by the job URL).

This script does not click Apply or submit anything. It only builds
the list so you can act on it quickly.
"""

import json
import re
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG - edit this section to change what counts as a match
# ---------------------------------------------------------------------------

TRACKER_PATH = Path("tracker.xlsx")

# Keywords that must appear (case-insensitive) in the title for a posting
# to be considered at all.
TITLE_MUST_INCLUDE = [
    "devops", "dev ops", "cloud engineer", "platform engineer",
    "site reliability", "sre", "infrastructure engineer",
    "cloud infrastructure", "kubernetes", "gitops",
]

# Titles containing any of these are dropped, even if they matched above -
# they're clearly not junior/entry roles.
TITLE_EXCLUDE = [
    "senior", "sr.", "sr ", "lead", "principal", "staff", "architect",
    "manager", "head of", "director", "chief",
]

# Words that, if present, are a good sign the role is junior/entry-level.
# Not required - many genuinely junior roles just say "DevOps Engineer"
# with no level word at all - but used to boost the match score.
JUNIOR_HINTS = [
    "junior", "entry", "entry-level", "entry level", "fresh grad",
    "fresh graduate", "0-2 years", "0-1 years", "1-2 years", "intern",
    "graduate program", "trainee",
]

# Her skills, pulled from the CV, used to score how well a posting matches.
CV_SKILLS = [
    "aws", "ec2", "s3", "emr", "docker", "terraform", "ci/cd", "kubernetes",
    "prometheus", "grafana", "linux", "git", "github actions", "jenkins",
    "python", "pyspark", "mongodb", "sql", "argocd", "gitops",
    "kustomize", "argo rollouts", "sbom", "syft", "k3d", "langchain",
    "huggingface", "rag", "node.js", "react", "typescript",
]

# Location filter for Wuzzuf results: keep roles based in these cities,
# plus anything explicitly marked remote.
CAIRO_AREA_HINTS = [
    "cairo", "giza", "nasr city", "maadi", "zamalek", "mohandessin",
    "heliopolis", "dokki", "new cairo", "6th of october", "sheikh zayed",
    "smart village", "downtown", "sheraton",
]

WUZZUF_RSS_URL = "https://wuzzuf.net/feeds/all-jobs.xml"
REMOTEOK_API_URL = "https://remoteok.com/api"

HEADERS = {"User-Agent": "Mozilla/5.0 (job-digest-bot; personal use)"}


# ---------------------------------------------------------------------------
# Fetch helpers
# ---------------------------------------------------------------------------

def fetch_url(url: str) -> bytes:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def matches_title(title: str) -> bool:
    t = title.lower()
    if not any(k in t for k in TITLE_MUST_INCLUDE):
        return False
    if any(k in t for k in TITLE_EXCLUDE):
        return False
    return True


def score_match(title: str, description: str) -> tuple[int, str]:
    """Returns (score, reason) - higher score = better fit for the CV."""
    text = f"{title} {description}".lower()
    hit_skills = [s for s in CV_SKILLS if s in text]
    score = len(hit_skills)
    if any(h in text for h in JUNIOR_HINTS):
        score += 3
    reason = ", ".join(hit_skills[:6]) if hit_skills else "title keyword match only"
    return score, reason


# ---------------------------------------------------------------------------
# Source: Wuzzuf RSS
# ---------------------------------------------------------------------------

def fetch_wuzzuf_jobs() -> list[dict]:
    jobs = []
    try:
        raw = fetch_url(WUZZUF_RSS_URL)
        root = ET.fromstring(raw)
    except Exception as exc:  # network hiccup, malformed feed, etc.
        print(f"[wuzzuf] skipped this run: {exc}")
        return jobs

    for item in root.iter("item"):
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        description = (item.findtext("description") or "").strip()
        pub_date = (item.findtext("pubDate") or "").strip()

        if not title or not link:
            continue
        if not matches_title(title):
            continue

        loc_text = f"{title} {description}".lower()
        is_cairo_area = any(h in loc_text for h in CAIRO_AREA_HINTS)
        is_remote = "remote" in loc_text
        if not (is_cairo_area or is_remote):
            continue

        score, reason = score_match(title, description)
        jobs.append({
            "source": "Wuzzuf",
            "title": title,
            "company": "See listing",
            "location": "Cairo/Giza area" if is_cairo_area else "Remote (verify region)",
            "posted": pub_date or "Unknown",
            "url": link,
            "match_score": score,
            "why_it_fits": reason,
        })
    return jobs


# ---------------------------------------------------------------------------
# Source: RemoteOK
# ---------------------------------------------------------------------------

def fetch_remoteok_jobs() -> list[dict]:
    jobs = []
    try:
        raw = fetch_url(REMOTEOK_API_URL)
        data = json.loads(raw)
    except Exception as exc:
        print(f"[remoteok] skipped this run: {exc}")
        return jobs

    for entry in data:
        if not isinstance(entry, dict) or "position" not in entry:
            continue  # first element is metadata, not a job
        title = entry.get("position", "")
        if not matches_title(title):
            continue

        description = entry.get("description", "") or ""
        tags = " ".join(entry.get("tags", []) or [])
        company = entry.get("company", "Unknown company")
        url = entry.get("url", "") or f"https://remoteok.com/remote-jobs/{entry.get('id','')}"
        posted = entry.get("date", "Unknown")

        score, reason = score_match(title, f"{description} {tags}")
        jobs.append({
            "source": "RemoteOK",
            "title": title,
            "company": company,
            # RemoteOK is global - flagged so you double check it's actually
            # open to Egypt/EMEA without sponsorship before applying.
            "location": "Remote (verify region/visa terms on listing)",
            "posted": posted,
            "url": url,
            "match_score": score,
            "why_it_fits": reason,
        })
    return jobs


# ---------------------------------------------------------------------------
# Tracker sheet read/write
# ---------------------------------------------------------------------------

HEADER = [
    "Date Found", "Match Score", "Why It Fits", "Title", "Company",
    "Location", "Posted", "Source", "Apply Link", "Applied?",
]

APPLIED_CHOICES = ["No", "Applied", "Interviewing", "Rejected", "Not Interested"]


def load_existing(path: Path) -> dict:
    """Returns {url: {"applied": str, "date_found": str}} for rows already tracked."""
    existing = {}
    if not path.exists():
        return existing
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    try:
        url_idx = header.index("Apply Link")
        applied_idx = header.index("Applied?")
        date_idx = header.index("Date Found")
    except ValueError:
        return existing  # unexpected sheet shape, start fresh
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= url_idx:
            continue
        url = row[url_idx]
        if not url:
            continue
        existing[url] = {
            "applied": row[applied_idx] or "No",
            "date_found": row[date_idx] or "",
        }
    return existing


def write_tracker(path: Path, all_jobs: list[dict], existing: dict) -> int:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    new_count = 0

    rows = []
    for job in all_jobs:
        url = job["url"]
        if url in existing:
            date_found = existing[url]["date_found"] or today
            applied = existing[url]["applied"]
        else:
            date_found = today
            applied = "No"
            new_count += 1
        rows.append([
            date_found, job["match_score"], job["why_it_fits"], job["title"],
            job["company"], job["location"], job["posted"], job["source"],
            url, applied,
        ])

    # Keep tracked jobs that this run didn't refetch (e.g. feed order changed),
    # so nothing you've already marked "Applied" silently disappears.
    seen_urls = {job["url"] for job in all_jobs}
    for url, info in existing.items():
        if url not in seen_urls:
            rows.append([
                info["date_found"], "", "(not refetched this run)", "", "",
                "", "", "", url, info["applied"],
            ])

    rows.sort(key=lambda r: (r[9] == "Applied", -(r[1] if isinstance(r[1], int) else 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Digest"

    ws.append(HEADER)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append(row)

    body_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [12, 11, 30, 32, 20, 24, 14, 10, 42, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(APPLIED_CHOICES)}"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"J2:J{ws.max_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return new_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    existing = load_existing(TRACKER_PATH)

    jobs = fetch_wuzzuf_jobs() + fetch_remoteok_jobs()

    # De-duplicate by URL within this run.
    dedup = {}
    for job in jobs:
        dedup[job["url"]] = job
    jobs = list(dedup.values())

    new_count = write_tracker(TRACKER_PATH, jobs, existing)
    print(f"Checked {len(jobs)} matching postings this run, {new_count} new since last run.")


if __name__ == "__main__":
    main()
